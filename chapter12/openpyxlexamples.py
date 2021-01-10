#! /usr/bin/env python3

import openpyxl
# Getting the sheets from the workbook
wb = openpyxl.load_workbook('./automate_online-materials/example.xlsx')
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet)
print(sheet.title)
anotherSheet = wb.get_active_sheet()
print(anotherSheet)

# Getting Cells from the Sheets
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)
print(sheet['C1'].value)

print(sheet.cell(row = 1, column = 2))
print(sheet.cell(row = 1, column = 2).value)
for i in range(1, 8 ,2):
    print(i, sheet.cell(row = i, column = 2).value)

